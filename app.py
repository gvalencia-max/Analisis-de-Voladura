"""
SPC Blasting Analysis — Streamlit Web Application
Kuz-Ram + Sadovsky + SPC + Monte Carlo
Universidad Nacional del Altiplano — Ing. de Minas
Autor: Giovany Valencia
"""

import io
import warnings
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib.gridspec import GridSpec
from scipy import stats
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import streamlit as st

warnings.filterwarnings("ignore")

# ─────────────────────────────────────────────
#  Matplotlib style
# ─────────────────────────────────────────────
plt.rcParams.update({
    "figure.dpi": 120,
    "font.family": "DejaVu Sans",
    "axes.titlesize": 11,
    "axes.labelsize": 10,
    "xtick.labelsize": 9,
    "ytick.labelsize": 9,
    "legend.fontsize": 9,
    "axes.grid": True,
    "grid.alpha": 0.3,
    "axes.spines.top": False,
    "axes.spines.right": False,
})

# ─────────────────────────────────────────────
#  Excel style constants
# ─────────────────────────────────────────────
HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT  = Font(color="FFFFFF", bold=True, name="Calibri", size=10)
SECTION_FILL = PatternFill("solid", fgColor="BDD7EE")
SECTION_FONT = Font(bold=True, name="Calibri", size=10)
GREEN_FILL   = PatternFill("solid", fgColor="C6EFCE")
YELLOW_FILL  = PatternFill("solid", fgColor="FFEB9C")
ORANGE_FILL  = PatternFill("solid", fgColor="FFCC99")
RED_FILL     = PatternFill("solid", fgColor="FFC7CE")
WHITE_FILL   = PatternFill("solid", fgColor="FFFFFF")
LIGHT_FILL   = PatternFill("solid", fgColor="F2F2F2")

# ─────────────────────────────────────────────
#  SPC constants table
# ─────────────────────────────────────────────
SPC_CONSTANTS = {
    2:  {"A2": 1.880, "D3": 0.000, "D4": 3.267, "d2": 1.128},
    3:  {"A2": 1.023, "D3": 0.000, "D4": 2.575, "d2": 1.693},
    4:  {"A2": 0.729, "D3": 0.000, "D4": 2.282, "d2": 2.059},
    5:  {"A2": 0.577, "D3": 0.000, "D4": 2.114, "d2": 2.326},
    6:  {"A2": 0.483, "D3": 0.000, "D4": 2.004, "d2": 2.534},
    7:  {"A2": 0.419, "D3": 0.076, "D4": 1.924, "d2": 2.704},
    8:  {"A2": 0.373, "D3": 0.136, "D4": 1.864, "d2": 2.847},
    10: {"A2": 0.308, "D3": 0.223, "D4": 1.777, "d2": 3.078},
}

PARAM_INFO = {
    "A":       {"label": "Rock Factor (Blastability)",              "symbol": "A",       "unit": "—",              "default": 8.0,   "std_frac": 0.15, "min": 0.6,   "max": 22.0,   "step": 0.1},
    "K":       {"label": "Powder Factor (specific charge)",         "symbol": "K",       "unit": "kg/m³",          "default": 0.40,  "std_frac": 0.12, "min": 0.05,  "max": 2.0,    "step": 0.01},
    "Q":       {"label": "Explosive charge per hole",               "symbol": "Q",       "unit": "kg",             "default": 150.0, "std_frac": 0.10, "min": 10.0,  "max": 1000.0, "step": 1.0},
    "RWS":     {"label": "Relative Weight Strength (ANFO=100)",     "symbol": "RWS",     "unit": "—",              "default": 105.0, "std_frac": 0.05, "min": 50.0,  "max": 150.0,  "step": 1.0},
    "d":       {"label": "Blasthole diameter",                      "symbol": "d",       "unit": "mm",             "default": 127.0, "std_frac": 0.0,  "min": 50.0,  "max": 400.0,  "step": 1.0},
    "B":       {"label": "Burden",                                  "symbol": "B",       "unit": "m",              "default": 4.5,   "std_frac": 0.08, "min": 0.5,   "max": 20.0,   "step": 0.1},
    "S":       {"label": "Spacing between holes",                   "symbol": "S",       "unit": "m",              "default": 5.0,   "std_frac": 0.08, "min": 0.5,   "max": 25.0,   "step": 0.1},
    "H":       {"label": "Bench height",                            "symbol": "H",       "unit": "m",              "default": 10.0,  "std_frac": 0.0,  "min": 1.0,   "max": 30.0,   "step": 0.5},
    "W":       {"label": "Drilling standard deviation",             "symbol": "W",       "unit": "m",              "default": 0.03,  "std_frac": 0.20, "min": 0.001, "max": 0.20,   "step": 0.001},
    "BCL":     {"label": "Bottom charge length",                    "symbol": "BCL",     "unit": "m",              "default": 3.0,   "std_frac": 0.0,  "min": 0.5,   "max": 10.0,   "step": 0.1},
    "CCL":     {"label": "Column charge length",                    "symbol": "CCL",     "unit": "m",              "default": 7.0,   "std_frac": 0.0,  "min": 0.5,   "max": 20.0,   "step": 0.1},
    "K_s":     {"label": "Site coefficient (Sadovsky PPV)",         "symbol": "K_s",     "unit": "—",              "default": 450.0, "std_frac": 0.25, "min": 50.0,  "max": 1500.0, "step": 10.0},
    "alpha":   {"label": "Attenuation exponent (Sadovsky PPV)",     "symbol": "α",       "unit": "—",              "default": 1.60,  "std_frac": 0.10, "min": 0.8,   "max": 2.5,    "step": 0.05},
    "R":       {"label": "Distance to control point",               "symbol": "R",       "unit": "m",              "default": 200.0, "std_frac": 0.0,  "min": 10.0,  "max": 5000.0, "step": 10.0},
    "Qd":      {"label": "Maximum charge per delay",                "symbol": "Qd",      "unit": "kg",             "default": 100.0, "std_frac": 0.10, "min": 5.0,   "max": 500.0,  "step": 5.0},
    "PPV_max": {"label": "Maximum allowable PPV",                   "symbol": "PPV_max", "unit": "mm/s",           "default": 25.0,  "std_frac": 0.0,  "min": 1.0,   "max": 200.0,  "step": 1.0},
    "X50_target": {"label": "Target X50 (design centre)",          "symbol": "X50_esp", "unit": "mm",             "default": 300.0, "std_frac": 0.0,  "min": 50.0,  "max": 1000.0, "step": 10.0},
    "X50_tol": {"label": "Tolerance on X50 (±value)",              "symbol": "±tol",    "unit": "mm",             "default": 80.0,  "std_frac": 0.0,  "min": 10.0,  "max": 300.0,  "step": 5.0},
}

# ─────────────────────────────────────────────
#  Calculation functions (unchanged from original)
# ─────────────────────────────────────────────
def calc_kuz_ram(A, K, Q, RWS, d, B, S, H, W, BCL, CCL):
    X50_cm = A * (K ** -0.8) * (Q ** (1.0/6.0)) * ((115.0/RWS) ** (19.0/20.0))
    X50 = X50_cm * 10.0
    L = BCL + CCL
    d_m = d / 1000.0
    n_u = ((2.2 - 14.0*B/d_m) * ((1.0+S/B)/2.0)**0.5
           * (1.0 - W/B) * (L/H)**0.3
           * (BCL/L + CCL/L))
    n_u = max(0.5, min(n_u, 2.5))
    Xc = X50 / (np.log(2) ** (1.0/n_u))
    return {"X50": X50, "n_uniformity": n_u, "Xc": Xc, "L": L}

def calc_rosin_rammler(x_arr, Xc, n_u):
    return 1.0 - np.exp(-(x_arr/Xc)**n_u)

def calc_ppv(K_s, alpha, R, Qd):
    SD = R / (Qd**0.5)
    PPV = K_s * (SD**(-alpha))
    return {"PPV": PPV, "SD": SD}

def calc_spc(subgroups, n_sub):
    means  = np.array([sg.mean() for sg in subgroups])
    ranges = np.array([sg.max() - sg.min() for sg in subgroups])
    X_bar_bar = means.mean()
    R_bar = ranges.mean()
    c = SPC_CONSTANTS[n_sub]
    A2, D3, D4, d2 = c["A2"], c["D3"], c["D4"], c["d2"]
    UCL_X = X_bar_bar + A2*R_bar
    LCL_X = X_bar_bar - A2*R_bar
    UCL_R = D4*R_bar
    LCL_R = D3*R_bar
    sigma_hat = R_bar / d2
    return {
        "means": means, "ranges": ranges,
        "X_bar_bar": X_bar_bar, "R_bar": R_bar,
        "UCL_X": UCL_X, "LCL_X": LCL_X, "CL_X": X_bar_bar,
        "UCL_R": UCL_R, "LCL_R": LCL_R, "CL_R": R_bar,
        "sigma_hat": sigma_hat, "A2": A2, "D3": D3, "D4": D4, "d2": d2,
    }

def calc_capability(X_bar_bar, sigma_hat, USL, LSL):
    Cp  = (USL - LSL) / (6.0*sigma_hat)
    Cpu = (USL - X_bar_bar) / (3.0*sigma_hat)
    Cpl = (X_bar_bar - LSL) / (3.0*sigma_hat)
    Cpk = min(Cpu, Cpl)
    return {"Cp": Cp, "Cpu": Cpu, "Cpl": Cpl, "Cpk": Cpk}

def nelson_rules(means, UCL_X, LCL_X, CL_X, sigma_hat):
    violations = {}
    n = len(means)
    r1 = [i for i in range(n) if means[i] > UCL_X or means[i] < LCL_X]
    if r1:
        violations[1] = r1
    for sign in [1, -1]:
        for i in range(n-8):
            if all((means[i+j] - CL_X)*sign > 0 for j in range(9)):
                violations.setdefault(2, []).extend(range(i, i+9))
    if 2 in violations:
        violations[2] = sorted(set(violations[2]))
    for i in range(n-5):
        seg = means[i:i+6]
        if all(seg[j] < seg[j+1] for j in range(5)) or all(seg[j] > seg[j+1] for j in range(5)):
            violations.setdefault(3, []).extend(range(i, i+6))
    if 3 in violations:
        violations[3] = sorted(set(violations[3]))
    return violations

def monte_carlo(p, n_mc):
    rng = np.random.default_rng(42)

    def tnorm(mean, std, low=None, high=None):
        if std == 0:
            return np.full(n_mc, mean)
        low  = low  if low  is not None else mean - 4*std
        high = high if high is not None else mean + 4*std
        a_s = (low - mean)/std
        b_s = (high - mean)/std
        return stats.truncnorm.rvs(a_s, b_s, loc=mean, scale=std, size=n_mc, random_state=rng)

    def lnorm(mean, cov):
        sigma2 = np.log(1 + cov**2)
        mu = np.log(mean) - sigma2/2
        return np.random.lognormal(mu, np.sqrt(sigma2), size=n_mc)

    A_mc     = tnorm(p["A"],    p["A"]*PARAM_INFO["A"]["std_frac"],      low=0.6,   high=22.0)
    K_mc     = tnorm(p["K"],    p["K"]*PARAM_INFO["K"]["std_frac"],      low=0.05,  high=2.0)
    Q_mc     = lnorm(p["Q"],    PARAM_INFO["Q"]["std_frac"])
    RWS_mc   = tnorm(p["RWS"],  p["RWS"]*PARAM_INFO["RWS"]["std_frac"],  low=50.0,  high=150.0)
    B_mc     = tnorm(p["B"],    p["B"]*PARAM_INFO["B"]["std_frac"],      low=0.5,   high=20.0)
    S_mc     = tnorm(p["S"],    p["S"]*PARAM_INFO["S"]["std_frac"],      low=0.5,   high=25.0)
    W_mc     = tnorm(p["W"],    p["W"]*PARAM_INFO["W"]["std_frac"],      low=0.001, high=0.2)
    K_s_mc   = lnorm(p["K_s"], PARAM_INFO["K_s"]["std_frac"])
    alpha_mc = tnorm(p["alpha"],p["alpha"]*PARAM_INFO["alpha"]["std_frac"], low=0.8, high=2.5)
    Qd_mc    = tnorm(p["Qd"],   p["Qd"]*PARAM_INFO["Qd"]["std_frac"],   low=5.0,   high=1000.0)

    d_mc   = np.full(n_mc, p["d"])
    H_mc   = np.full(n_mc, p["H"])
    BCL_mc = np.full(n_mc, p["BCL"])
    CCL_mc = np.full(n_mc, p["CCL"])
    R_mc   = np.full(n_mc, p["R"])

    L_mc    = BCL_mc + CCL_mc
    X50_mc  = A_mc*(K_mc**-0.8)*(Q_mc**(1.0/6.0))*((115.0/RWS_mc)**(19.0/20.0))*10.0
    d_m_mc  = d_mc/1000.0
    n_u_mc  = ((2.2 - 14.0*B_mc/d_m_mc)
                *((1.0+S_mc/B_mc)/2.0)**0.5
                *(1.0-W_mc/B_mc)
                *(L_mc/H_mc)**0.3
                *(BCL_mc/L_mc + CCL_mc/L_mc))
    n_u_mc  = np.clip(n_u_mc, 0.5, 2.5)
    SD_mc   = R_mc/(Qd_mc**0.5)
    PPV_mc  = K_s_mc*(SD_mc**(-alpha_mc))

    USL = p["X50_target"] + p["X50_tol"]
    LSL = p["X50_target"] - p["X50_tol"]
    pct_X50_ok = np.mean((X50_mc >= LSL) & (X50_mc <= USL))*100.0
    pct_PPV_ok = np.mean(PPV_mc <= p["PPV_max"])*100.0

    def stats_dict(arr):
        return {"mean": arr.mean(), "std": arr.std(),
                "P5": np.percentile(arr,5), "P50": np.percentile(arr,50),
                "P90": np.percentile(arr,90), "P95": np.percentile(arr,95)}

    input_arrays = {"A": A_mc,"K": K_mc,"Q": Q_mc,"RWS": RWS_mc,
                    "B": B_mc,"S": S_mc,"W": W_mc,"K_s": K_s_mc,"alpha": alpha_mc,"Qd": Qd_mc}
    corr_X50 = {k: np.corrcoef(v, X50_mc)[0,1] for k,v in input_arrays.items()}
    corr_PPV = {k: np.corrcoef(v, PPV_mc)[0,1] for k,v in input_arrays.items()}
    sample_size = min(500, n_mc)
    sample = {"A": A_mc[:sample_size],"K": K_mc[:sample_size],"Q": Q_mc[:sample_size],
              "RWS": RWS_mc[:sample_size],"B": B_mc[:sample_size],"S": S_mc[:sample_size],
              "W": W_mc[:sample_size],"K_s": K_s_mc[:sample_size],"alpha": alpha_mc[:sample_size],
              "Qd": Qd_mc[:sample_size],"X50": X50_mc[:sample_size],"PPV": PPV_mc[:sample_size],
              "n_uniformity": n_u_mc[:sample_size]}
    return {"X50": stats_dict(X50_mc),"PPV": stats_dict(PPV_mc),"n_u": stats_dict(n_u_mc),
            "pct_X50_ok": pct_X50_ok,"pct_PPV_ok": pct_PPV_ok,
            "corr_X50": corr_X50,"corr_PPV": corr_PPV,"sample": sample,
            "X50_mc": X50_mc,"PPV_mc": PPV_mc}

# ─────────────────────────────────────────────
#  Interpretation helpers
# ─────────────────────────────────────────────
def interpret_X50(X50, USL, LSL, target):
    if X50 > USL:
        return "OUT OF SPEC (above USL)", "Increase powder factor K or reduce Burden/Spacing. Consider a higher RWS explosive.", "red"
    elif X50 < LSL:
        return "OUT OF SPEC (below LSL)", "Reduce powder factor K or increase Burden/Spacing to avoid over-blasting.", "orange"
    elif X50 > target + 0.5*(USL-target):
        return "WITHIN SPEC — approaching upper limit", "Monitor closely. Small increase in K may be needed.", "yellow"
    elif X50 < target - 0.5*(target-LSL):
        return "WITHIN SPEC — approaching lower limit", "Slightly reduce explosive load or increase stemming.", "yellow"
    else:
        return "WITHIN SPEC — on target", "Maintain current blasting parameters. SPC monitoring recommended.", "green"

def interpret_PPV(PPV, PPV_max):
    ratio = PPV/PPV_max
    if ratio > 1.0:
        return "EXCEEDS LIMIT — structural damage risk", "Reduce Qd immediately. Increase number of delays.", "red"
    elif ratio > 0.75:
        return "WITHIN LIMIT — close to threshold", "Reduce Qd by 20-30%. Verify geotechnical conditions.", "orange"
    elif ratio > 0.50:
        return "ACCEPTABLE — monitor regularly", f"Safety factor ≈ {1/ratio:.1f}x is acceptable.", "yellow"
    else:
        return "SAFE — ample margin", f"Safety factor = {1/ratio:.1f}x. Qd can be increased if needed.", "green"

def interpret_capability(Cp, Cpk):
    if Cpk < 1.0:
        return "NOT CAPABLE — producing out-of-spec blasts", "Redesign blasting process: review explosive type, drilling precision, bench geometry.", "red"
    elif Cpk < 1.33:
        return "MARGINALLY CAPABLE — improve and monitor", "Increase drilling precision (reduce W), optimize K, tighten subgroup monitoring.", "orange"
    elif Cp >= 2.0 and Cpk >= 2.0:
        return "VERY CAPABLE (Six Sigma level)", "Optimize for cost reduction. Consider relaxing specifications.", "green"
    else:
        return "CAPABLE — maintain SPC control", "Keep SPC active. Centering improvement (reduce Cp-Cpk gap) possible.", "green"

# ─────────────────────────────────────────────
#  Plot functions — return matplotlib figures
# ─────────────────────────────────────────────
def fig_kuz_ram(kuz):
    x_arr = np.linspace(1, max(kuz["Xc"]*3, 1000), 500)
    passing = calc_rosin_rammler(x_arr, kuz["Xc"], kuz["n_uniformity"])
    fig, ax = plt.subplots(figsize=(8,5))
    ax.plot(x_arr, passing*100, "b-", lw=2, label=f"Rosin-Rammler (n={kuz['n_uniformity']:.2f})")
    ax.axvline(kuz["X50"], color="red", ls="--", lw=1.5, label=f"X50 = {kuz['X50']:.0f} mm")
    ax.axhline(50, color="gray", ls=":", lw=1)
    ax.set_xlabel("Fragment size x (mm)")
    ax.set_ylabel("Cumulative passing (%)")
    ax.set_title("Kuz-Ram Particle Size Distribution (Rosin-Rammler)")
    ax.legend(); ax.set_xscale("log"); ax.set_xlim(left=1); ax.set_ylim(0,100)
    fig.tight_layout()
    return fig

def fig_ppv(p, ppv_res):
    SD_arr = np.linspace(1, 150, 400)
    PPV_arr = p["K_s"]*(SD_arr**(-p["alpha"]))
    fig, ax = plt.subplots(figsize=(8,5))
    ax.loglog(SD_arr, PPV_arr, "b-", lw=2, label="Sadovsky curve")
    ax.axhline(p["PPV_max"], color="red", ls="--", lw=1.5, label=f"PPV_max = {p['PPV_max']} mm/s")
    ax.scatter([ppv_res["SD"]], [ppv_res["PPV"]], color="orange", s=80, zorder=5,
               label=f"Design: SD={ppv_res['SD']:.1f}, PPV={ppv_res['PPV']:.2f} mm/s")
    ax.set_xlabel("Scaled Distance SD = R / √Qd  (m/kg⁰·⁵)")
    ax.set_ylabel("Peak Particle Velocity PPV (mm/s)")
    ax.set_title("Sadovsky PPV Attenuation Curve")
    ax.legend()
    fig.tight_layout()
    return fig

def fig_xbar(spc, p):
    means = spc["means"]; x = np.arange(1, len(means)+1)
    USL = p["X50_target"]+p["X50_tol"]; LSL = p["X50_target"]-p["X50_tol"]
    fig, (ax1, ax2) = plt.subplots(2,1, figsize=(10,7), sharex=True)
    for ax, data, ucl, lcl, cl, ylabel, title in [
        (ax1, means, spc["UCL_X"], spc["LCL_X"], spc["CL_X"], "X̄ (mm)", "X̄ Chart — Subgroup Means"),
        (ax2, spc["ranges"], spc["UCL_R"], spc["LCL_R"], spc["CL_R"], "R (mm)", "R Chart — Subgroup Ranges"),
    ]:
        ax.plot(x, data, "b-o", lw=1.5, ms=6, label="Data")
        ax.axhline(ucl, color="red", ls="--", lw=1.5, label=f"UCL={ucl:.1f}")
        ax.axhline(lcl, color="red", ls="--", lw=1.5, label=f"LCL={lcl:.1f}")
        ax.axhline(cl,  color="green", ls="-", lw=1.5, label=f"CL={cl:.1f}")
        out = [i for i,v in enumerate(data) if v>ucl or v<lcl]
        if out:
            ax.scatter([x[i] for i in out],[data[i] for i in out], color="red", s=100, zorder=6, label="Out of control")
        ax.set_ylabel(ylabel); ax.set_title(title); ax.legend(loc="upper right", fontsize=8)
    ax2.set_xlabel("Subgroup number"); ax2.set_xticks(x)
    fig.tight_layout()
    return fig

def fig_capability(spc, p, cap):
    USL = p["X50_target"]+p["X50_tol"]; LSL = p["X50_target"]-p["X50_tol"]
    mu = spc["X_bar_bar"]; sigma = spc["sigma_hat"]
    x = np.linspace(mu-5*sigma, mu+5*sigma, 500)
    y = stats.norm.pdf(x, mu, sigma)
    fig, ax = plt.subplots(figsize=(9,5))
    ax.plot(x, y, "b-", lw=2, label=f"Process N({mu:.1f}, {sigma:.1f}²)")
    ax.fill_between(x, y, where=(x>=LSL)&(x<=USL), alpha=0.25, color="green", label="Within spec")
    ax.fill_between(x, y, where=(x<LSL),  alpha=0.35, color="red", label="Below LSL")
    ax.fill_between(x, y, where=(x>USL),  alpha=0.35, color="red", label="Above USL")
    ax.axvline(USL, color="red",   ls="--", lw=1.5, label=f"USL={USL:.0f}")
    ax.axvline(LSL, color="red",   ls="--", lw=1.5, label=f"LSL={LSL:.0f}")
    ax.axvline(mu,  color="blue",  ls="-",  lw=1.2, label=f"X̄={mu:.1f}")
    ax.axvline(p["X50_target"], color="green", ls=":", lw=1.2, label=f"Target={p['X50_target']:.0f}")
    ax.set_xlabel("X50 (mm)"); ax.set_ylabel("Probability Density")
    ax.set_title(f"Process Capability — Cp={cap['Cp']:.2f}, Cpk={cap['Cpk']:.2f}")
    ax.legend(fontsize=8)
    fig.tight_layout()
    return fig

def fig_mc_x50(mc, p):
    USL = p["X50_target"]+p["X50_tol"]; LSL = p["X50_target"]-p["X50_tol"]
    fig, ax = plt.subplots(figsize=(9,5))
    ax.hist(mc["X50_mc"], bins=80, color="steelblue", edgecolor="white", alpha=0.8, density=True)
    ax.axvline(USL, color="red", ls="--", lw=1.5, label=f"USL={USL:.0f}")
    ax.axvline(LSL, color="red", ls="--", lw=1.5, label=f"LSL={LSL:.0f}")
    ax.axvline(mc["X50"]["P50"], color="orange", ls="-", lw=1.5, label=f"P50={mc['X50']['P50']:.0f}")
    ax.axvline(mc["X50"]["mean"], color="blue", ls="-", lw=1.5, label=f"Mean={mc['X50']['mean']:.0f}")
    ax.set_xlabel("X50 (mm)"); ax.set_ylabel("Density")
    ax.set_title(f"Monte Carlo — X50 Distribution  |  In-spec: {mc['pct_X50_ok']:.1f}%")
    ax.legend(fontsize=8)
    fig.tight_layout()
    return fig

def fig_mc_ppv(mc, p):
    fig, ax = plt.subplots(figsize=(9,5))
    ax.hist(mc["PPV_mc"], bins=80, color="coral", edgecolor="white", alpha=0.8, density=True)
    ax.axvline(p["PPV_max"], color="red", ls="--", lw=1.5, label=f"PPV_max={p['PPV_max']}")
    ax.axvline(mc["PPV"]["P95"], color="purple", ls="-", lw=1.5, label=f"P95={mc['PPV']['P95']:.2f}")
    ax.axvline(mc["PPV"]["mean"], color="blue", ls="-", lw=1.5, label=f"Mean={mc['PPV']['mean']:.2f}")
    ax.set_xlabel("PPV (mm/s)"); ax.set_ylabel("Density")
    ax.set_title(f"Monte Carlo — PPV Distribution  |  Within limit: {mc['pct_PPV_ok']:.1f}%")
    ax.legend(fontsize=8)
    fig.tight_layout()
    return fig

def fig_tornado(mc):
    corr = mc["corr_X50"]
    keys = sorted(corr, key=lambda k: abs(corr[k]), reverse=True)
    vals = [corr[k] for k in keys]
    colors = ["steelblue" if v>=0 else "salmon" for v in vals]
    fig, ax = plt.subplots(figsize=(8,5))
    ax.barh(range(len(keys)), vals, color=colors, edgecolor="white")
    ax.set_yticks(range(len(keys))); ax.set_yticklabels(keys)
    ax.axvline(0, color="black", lw=0.8)
    ax.set_xlabel("Pearson Correlation with X50"); ax.set_title("Tornado Chart — Sensitivity on X50")
    fig.tight_layout()
    return fig

def fig_tornado_ppv(mc):
    corr = mc["corr_PPV"]
    keys = sorted(corr, key=lambda k: abs(corr[k]), reverse=True)
    vals = [corr[k] for k in keys]
    colors = ["steelblue" if v>=0 else "salmon" for v in vals]
    fig, ax = plt.subplots(figsize=(8,5))
    ax.barh(range(len(keys)), vals, color=colors, edgecolor="white")
    ax.set_yticks(range(len(keys))); ax.set_yticklabels(keys)
    ax.axvline(0, color="black", lw=0.8)
    ax.set_xlabel("Pearson Correlation with PPV"); ax.set_title("Tornado Chart — Sensitivity on PPV")
    fig.tight_layout()
    return fig

# ─────────────────────────────────────────────
#  Excel export — returns bytes
# ─────────────────────────────────────────────
def _thin_border():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)

def _write_header_row(ws, r, cols):
    for j, v in enumerate(cols):
        c = ws.cell(row=r, column=1+j, value=v)
        c.fill=HEADER_FILL; c.font=HEADER_FONT
        c.alignment=Alignment(horizontal="center"); c.border=_thin_border()

def _write_section_row(ws, r, title, ncols=5):
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=ncols)
    c = ws.cell(row=r,column=1,value=title)
    c.fill=SECTION_FILL; c.font=SECTION_FONT; c.border=_thin_border()

def _write_data_row(ws, r, vals, fill=None):
    for j,v in enumerate(vals):
        c = ws.cell(row=r,column=1+j,value=v)
        c.border=_thin_border()
        if fill:
            c.fill=fill

def _auto_width(ws):
    for col in ws.columns:
        mx=0
        for cell in col:
            try:
                mx = max(mx, len(str(cell.value or "")))
            except:
                pass
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(mx+4, 40)

def _color_fill(color):
    return {"red": RED_FILL,"orange": ORANGE_FILL,"yellow": YELLOW_FILL,"green": GREEN_FILL}.get(color, WHITE_FILL)

def export_excel(p, kuz, ppv_res, spc, cap, mc_results, violations):
    wb = Workbook()
    USL = p["X50_target"]+p["X50_tol"]; LSL = p["X50_target"]-p["X50_tol"]

    # Sheet 1 – Input parameters
    ws1 = wb.active; ws1.title = "1_Input_Parameters"; ws1.sheet_view.showGridLines = False
    _write_header_row(ws1,1,["Section","Parameter","Symbol","Value","Units"])
    sections = [
        ("Kuz-Ram",["A","K","Q","RWS","d","B","S","H","W","BCL","CCL"]),
        ("Sadovsky PPV",["K_s","alpha","R","Qd","PPV_max"]),
        ("SPC & Specs",["X50_target","X50_tol"]),
    ]
    r=2
    for sec_name, keys in sections:
        _write_section_row(ws1,r,f"─── {sec_name} ───",ncols=5); r+=1
        for k in keys:
            info = PARAM_INFO.get(k,{})
            _write_data_row(ws1,r,[sec_name, info.get("label",k), info.get("symbol",k), p.get(k,""), info.get("unit","")]); r+=1
    _write_section_row(ws1,r,"─── SPC SUBGROUPS ───",ncols=5); r+=1
    for i,sg in enumerate(p["subgroups"]):
        _write_data_row(ws1,r,[f"Subgroup {i+1}", str(list(sg.round(1))), "", "", "mm"]); r+=1
    _auto_width(ws1)

    # Sheet 2 – Summary
    ws2 = wb.create_sheet("2_Summary_Results"); ws2.sheet_view.showGridLines=False
    _write_header_row(ws2,1,["Analysis","Result","Value","Interpretation","Recommendation"])
    x50_lvl,x50_rec,x50_col = interpret_X50(kuz["X50"],USL,LSL,p["X50_target"])
    ppv_lvl,ppv_rec,ppv_col = interpret_PPV(ppv_res["PPV"],p["PPV_max"])
    cap_lvl,cap_rec,cap_col = interpret_capability(cap["Cp"],cap["Cpk"])
    det_rows=[
        ("Kuz-Ram","X50",f"{kuz['X50']:.1f} mm",x50_lvl,x50_rec,x50_col),
        ("Kuz-Ram","n uniformity",f"{kuz['n_uniformity']:.3f}","—","—","green"),
        ("Sadovsky","PPV",f"{ppv_res['PPV']:.3f} mm/s",ppv_lvl,ppv_rec,ppv_col),
        ("SPC","Cp",f"{cap['Cp']:.2f}",cap_lvl,cap_rec,cap_col),
        ("SPC","Cpk",f"{cap['Cpk']:.2f}",cap_lvl,cap_rec,cap_col),
    ]
    mc_rows=[
        ("MC X50","Mean",f"{mc_results['X50']['mean']:.0f} mm",f"In-spec: {mc_results['pct_X50_ok']:.1f}%","—","green" if mc_results['pct_X50_ok']>80 else "yellow"),
        ("MC PPV","P95", f"{mc_results['PPV']['P95']:.2f} mm/s",f"Within limit: {mc_results['pct_PPV_ok']:.1f}%","—","green" if mc_results['pct_PPV_ok']>80 else "red"),
    ]
    r=2
    _write_section_row(ws2,r,"─── DETERMINISTIC RESULTS ───",ncols=5); r+=1
    for row_data in det_rows:
        a,res,val,interp,rec,color=row_data
        _write_data_row(ws2,r,[a,res,val,interp,rec])
        for j in range(5): ws2.cell(row=r,column=1+j).fill=_color_fill(color)
        r+=1
    _write_section_row(ws2,r,"─── MONTE CARLO PROBABILISTIC RESULTS ───",ncols=5); r+=1
    for row_data in mc_rows:
        a,res,val,interp,rec,color=row_data
        _write_data_row(ws2,r,[a,res,val,interp,rec])
        for j in range(5): ws2.cell(row=r,column=1+j).fill=_color_fill(color)
        r+=1
    _auto_width(ws2)

    # Sheet 3 – SPC calculations
    ws3 = wb.create_sheet("3_SPC_Calculations"); ws3.sheet_view.showGridLines=False
    _write_header_row(ws3,1,["Subgroup","Vol1","Vol2","Vol3","Vol4","Vol5","Mean X̄i (mm)","Range Ri (mm)","Status"])
    for i,sg in enumerate(p["subgroups"]):
        row_vals=[i+1]+list(sg)+[""]*(5-len(sg))
        row_vals.append(round(spc["means"][i],2)); row_vals.append(round(spc["ranges"][i],2))
        out = spc["means"][i]>spc["UCL_X"] or spc["means"][i]<spc["LCL_X"]
        row_vals.append("OUT OF CONTROL ⚠" if out else "In Control ✓")
        _write_data_row(ws3,2+i,row_vals, fill=RED_FILL if out else GREEN_FILL)
    _auto_width(ws3)

    # Sheet 4 – Monte Carlo stats
    ws4 = wb.create_sheet("4_MonteCarlo_Stats"); ws4.sheet_view.showGridLines=False
    _write_header_row(ws4,1,["Variable","Mean","Std Dev","P5","P50","P90","P95","Units"])
    r=2
    for label, st, unit in [("X50 (Kuz-Ram)",mc_results["X50"],"mm"),("PPV (Sadovsky)",mc_results["PPV"],"mm/s"),("n index",mc_results["n_u"],"—")]:
        _write_data_row(ws4,r,[label,round(st["mean"],4),round(st["std"],4),round(st["P5"],4),round(st["P50"],4),round(st["P90"],4),round(st["P95"],4),unit]); r+=1
    _auto_width(ws4)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

# ─────────────────────────────────────────────
#  Streamlit UI
# ─────────────────────────────────────────────
def main():
    st.set_page_config(
        page_title="SPC Blasting Analysis",
        page_icon="💥",
        layout="wide",
    )

    # ── Header ──────────────────────────────
    st.title("💥 SPC Blasting Analysis")
    st.markdown(
        "**Kuz-Ram Fragmentation · Sadovsky PPV · Statistical Process Control · Monte Carlo**  \n"
        "_Universidad Nacional del Altiplano — Ingeniería de Minas_"
    )
    st.divider()

    # ── Sidebar: all input parameters ───────
    with st.sidebar:
        st.header("⚙️ Input Parameters")

        st.subheader("Section 1 — Kuz-Ram Model")
        A   = st.number_input("Rock Factor A  [—]",       min_value=0.6,  max_value=22.0,  value=8.0,   step=0.1,  help="Very soft=1 · Medium=4.5 · Hard=10 · Very hard=18")
        K   = st.number_input("Powder Factor K  [kg/m³]", min_value=0.05, max_value=2.0,   value=0.40,  step=0.01, help="Light=0.1-0.25 · Typical=0.25-0.50 · Hard rock=0.50-0.80")
        Q   = st.number_input("Charge per hole Q  [kg]",  min_value=10.0, max_value=1000.0,value=150.0, step=1.0,  help="Small holes=50-100 · Medium=100-200 · Large=200-500")
        RWS = st.number_input("RWS (ANFO=100)  [—]",      min_value=50.0, max_value=150.0, value=105.0, step=1.0,  help="ANFO=100 · Heavy ANFO=110-120 · Emulsions=115-125")
        d   = st.number_input("Blasthole diameter d  [mm]",min_value=50.0, max_value=400.0, value=127.0, step=1.0,  help="Small=76-102 · Medium=115-165 · Large=200-311")
        B   = st.number_input("Burden B  [m]",            min_value=0.5,  max_value=20.0,  value=4.5,   step=0.1,  help="Small=2-3 m · Medium=3-5 m · Large=5-8 m")
        S   = st.number_input("Spacing S  [m]",           min_value=0.5,  max_value=25.0,  value=5.0,   step=0.1,  help="S/B ratio 1.0 (square) to 1.5 (wide)")
        H   = st.number_input("Bench height H  [m]",      min_value=1.0,  max_value=30.0,  value=10.0,  step=0.5,  help="Small=5-8 m · Typical=8-15 m · Large=15-20 m")
        W   = st.number_input("Drilling σ W  [m]",        min_value=0.001,max_value=0.20,  value=0.03,  step=0.001,help="Precision=0.01-0.02 · Typical=0.02-0.04 · Poor=0.04-0.06")
        BCL = st.number_input("Bottom charge BCL  [m]",   min_value=0.5,  max_value=10.0,  value=3.0,   step=0.1,  help="Small holes=1-2 m · Medium=2-3.5 m · Large=3-5 m")
        CCL = st.number_input("Column charge CCL  [m]",   min_value=0.5,  max_value=20.0,  value=7.0,   step=0.1,  help="Short=2-5 m · Typical=5-10 m · Long=8-15 m")

        st.subheader("Section 2 — Sadovsky PPV")
        K_s     = st.number_input("Site coeff K_s  [—]",       min_value=50.0,  max_value=1500.0, value=450.0, step=10.0, help="Hard rock=100-300 · Medium=200-500 · Soil=600-1200")
        alpha   = st.number_input("Attenuation α  [—]",        min_value=0.8,   max_value=2.5,    value=1.60,  step=0.05, help="Hard rock=1.7-2.0 · Medium=1.5-1.8 · Soil=1.0-1.3")
        R       = st.number_input("Distance R  [m]",           min_value=10.0,  max_value=5000.0, value=200.0, step=10.0, help="Near structures=50-100 m · Typical=100-500 m")
        Qd      = st.number_input("Max charge/delay Qd  [kg]", min_value=5.0,   max_value=500.0,  value=100.0, step=5.0,  help="Small=20-60 kg · Typical=60-150 kg · Large=150-300 kg")
        PPV_max = st.number_input("PPV limit  [mm/s]",         min_value=1.0,   max_value=200.0,  value=25.0,  step=1.0,  help="Historic bldg=3 · Residential NTP Peru=15 · Industrial=50")

        st.subheader("Section 3 — SPC Specifications")
        X50_target = st.number_input("Target X50  [mm]", min_value=50.0,  max_value=1000.0, value=300.0, step=10.0)
        X50_tol    = st.number_input("Tolerance ±  [mm]",min_value=10.0,  max_value=300.0,  value=80.0,  step=5.0)
        n_subgroup = st.selectbox("Subgroup size n", options=list(SPC_CONSTANTS.keys()), index=3)

        st.subheader("Section 4 — Monte Carlo")
        n_mc = st.select_slider("Simulations", options=[1000,5000,10000,25000,50000], value=10000)

    # ── Historical X50 data ──────────────────
    st.header("📋 Historical X50 Measurements (mm)")
    st.markdown(f"Enter your X50 data below. Each row = one subgroup of **n={n_subgroup}** blasts.")

    use_defaults = st.checkbox("Use default example dataset", value=True)

    if use_defaults:
        default_raw = [
            [320, 295, 340, 285, 310],
            [305, 280, 325, 300, 290],
            [350, 330, 360, 345, 340],
            [270, 295, 280, 300, 285],
            [315, 330, 305, 320, 340],
        ]
        default_raw = [d[:n_subgroup] for d in default_raw]
        subgroups = [np.array(row, dtype=float) for row in default_raw]
        st.info(f"Using 5-subgroup default dataset (n={n_subgroup} each).")
    else:
        n_groups = st.number_input("Number of subgroups", min_value=3, max_value=30, value=5, step=1)
        subgroups = []
        cols_sg = st.columns(min(n_subgroup, 5))
        for i in range(n_groups):
            st.markdown(f"**Subgroup {i+1}**")
            row_cols = st.columns(n_subgroup)
            row = []
            for j in range(n_subgroup):
                val = row_cols[j].number_input(
                    f"SG{i+1}-V{j+1}", min_value=0.0, max_value=2000.0,
                    value=300.0 + float(np.random.randint(-50,50)),
                    key=f"sg_{i}_{j}", label_visibility="collapsed"
                )
                row.append(val)
            subgroups.append(np.array(row, dtype=float))

    # ── Run analysis ─────────────────────────
    st.divider()
    run_btn = st.button("▶ Run Analysis", type="primary", use_container_width=True)

    if run_btn:
        p = {
            "A": A, "K": K, "Q": Q, "RWS": RWS, "d": d,
            "B": B, "S": S, "H": H, "W": W, "BCL": BCL, "CCL": CCL,
            "K_s": K_s, "alpha": alpha, "R": R, "Qd": Qd, "PPV_max": PPV_max,
            "X50_target": X50_target, "X50_tol": X50_tol,
            "n_subgroup": n_subgroup, "n_mc": n_mc,
            "subgroups": subgroups,
        }

        with st.spinner("Running analysis…"):
            kuz       = calc_kuz_ram(A, K, Q, RWS, d, B, S, H, W, BCL, CCL)
            ppv_res   = calc_ppv(K_s, alpha, R, Qd)
            spc       = calc_spc(subgroups, n_subgroup)
            USL = X50_target + X50_tol; LSL = X50_target - X50_tol
            cap       = calc_capability(spc["X_bar_bar"], spc["sigma_hat"], USL, LSL)
            violations= nelson_rules(spc["means"], spc["UCL_X"], spc["LCL_X"], spc["CL_X"], spc["sigma_hat"])
            mc        = monte_carlo(p, n_mc)

        # ── KPI Cards ────────────────────────
        st.header("📊 Key Results")
        c1, c2, c3, c4, c5 = st.columns(5)
        x50_lvl,_,x50_col = interpret_X50(kuz["X50"], USL, LSL, X50_target)
        ppv_lvl,_,ppv_col = interpret_PPV(ppv_res["PPV"], PPV_max)
        _,_,cap_col       = interpret_capability(cap["Cp"], cap["Cpk"])

        COLOR_EMOJI = {"green":"🟢","yellow":"🟡","orange":"🟠","red":"🔴"}
        c1.metric("X50 (Kuz-Ram)", f"{kuz['X50']:.0f} mm",  f"{COLOR_EMOJI[x50_col]} {x50_lvl[:12]}")
        c2.metric("PPV (Sadovsky)", f"{ppv_res['PPV']:.2f} mm/s", f"{COLOR_EMOJI[ppv_col]} {ppv_lvl[:12]}")
        c3.metric("Cp",  f"{cap['Cp']:.3f}",  f"{COLOR_EMOJI[cap_col]} Capability")
        c4.metric("Cpk", f"{cap['Cpk']:.3f}", "")
        c5.metric("n uniformity", f"{kuz['n_uniformity']:.3f}", "")

        # ── Tabs for each section ────────────
        tab1, tab2, tab3, tab4, tab5 = st.tabs(
            ["🪨 Kuz-Ram", "📳 Sadovsky PPV", "📈 SPC Charts", "🎲 Monte Carlo", "📥 Export"]
        )

        with tab1:
            st.subheader("Kuz-Ram Fragmentation Model")
            col_a, col_b = st.columns([2,1])
            with col_a:
                st.pyplot(fig_kuz_ram(kuz))
            with col_b:
                st.markdown(f"**X50** = {kuz['X50']:.1f} mm")
                st.markdown(f"**Xc** = {kuz['Xc']:.1f} mm")
                st.markdown(f"**n uniformity** = {kuz['n_uniformity']:.3f}")
                st.markdown(f"**Total charge length L** = {kuz['L']:.2f} m")
                st.markdown(f"**USL** = {USL:.0f} mm &nbsp; | &nbsp; **LSL** = {LSL:.0f} mm")
                lvl, rec, col = interpret_X50(kuz["X50"], USL, LSL, X50_target)
                if col == "green":   st.success(f"**{lvl}**\n\n{rec}")
                elif col == "yellow": st.warning(f"**{lvl}**\n\n{rec}")
                else:                 st.error(f"**{lvl}**\n\n{rec}")

        with tab2:
            st.subheader("Sadovsky PPV Attenuation")
            col_a, col_b = st.columns([2,1])
            with col_a:
                st.pyplot(fig_ppv(p, ppv_res))
            with col_b:
                st.markdown(f"**SD** = {ppv_res['SD']:.2f} m/kg⁰·⁵")
                st.markdown(f"**PPV** = {ppv_res['PPV']:.3f} mm/s")
                st.markdown(f"**PPV limit** = {PPV_max:.1f} mm/s")
                lvl, rec, col = interpret_PPV(ppv_res["PPV"], PPV_max)
                if col == "green":   st.success(f"**{lvl}**\n\n{rec}")
                elif col == "yellow": st.warning(f"**{lvl}**\n\n{rec}")
                else:                 st.error(f"**{lvl}**\n\n{rec}")

        with tab3:
            st.subheader("Statistical Process Control (X̄-R Charts)")
            st.pyplot(fig_xbar(spc, p))
            st.subheader("Process Capability")
            st.pyplot(fig_capability(spc, p, cap))

            col_a, col_b, col_c = st.columns(3)
            col_a.metric("X̄̄ (grand mean)", f"{spc['X_bar_bar']:.1f} mm")
            col_b.metric("σ̂ (estimated)",   f"{spc['sigma_hat']:.2f} mm")
            col_c.metric("R̄ (mean range)",   f"{spc['R_bar']:.1f} mm")

            if violations:
                st.warning("⚠️ **Nelson Rule Violations Detected:**")
                for rule, idxs in violations.items():
                    st.write(f"- Rule {rule}: subgroups {[i+1 for i in idxs]}")
            else:
                st.success("✅ No Nelson rule violations detected.")

            lvl, rec, col = interpret_capability(cap["Cp"], cap["Cpk"])
            if col == "green":   st.success(f"**{lvl}** — {rec}")
            elif col == "orange": st.warning(f"**{lvl}** — {rec}")
            else:                 st.error(f"**{lvl}** — {rec}")

        with tab4:
            st.subheader(f"Monte Carlo Simulation ({n_mc:,} runs)")
            col_a, col_b = st.columns(2)
            with col_a:
                st.pyplot(fig_mc_x50(mc, p))
            with col_b:
                st.pyplot(fig_mc_ppv(mc, p))

            col_a, col_b = st.columns(2)
            with col_a:
                st.pyplot(fig_tornado(mc))
            with col_b:
                st.pyplot(fig_tornado_ppv(mc))

            st.subheader("Monte Carlo Statistics")
            import pandas as pd
            mc_df = pd.DataFrame({
                "Variable": ["X50 (mm)", "PPV (mm/s)", "n uniformity"],
                "Mean":  [mc["X50"]["mean"], mc["PPV"]["mean"], mc["n_u"]["mean"]],
                "Std":   [mc["X50"]["std"],  mc["PPV"]["std"],  mc["n_u"]["std"]],
                "P5":    [mc["X50"]["P5"],   mc["PPV"]["P5"],   mc["n_u"]["P5"]],
                "P50":   [mc["X50"]["P50"],  mc["PPV"]["P50"],  mc["n_u"]["P50"]],
                "P90":   [mc["X50"]["P90"],  mc["PPV"]["P90"],  mc["n_u"]["P90"]],
                "P95":   [mc["X50"]["P95"],  mc["PPV"]["P95"],  mc["n_u"]["P95"]],
            }).set_index("Variable").round(3)
            st.dataframe(mc_df, use_container_width=True)

            col_a, col_b = st.columns(2)
            col_a.metric("X50 in-spec probability", f"{mc['pct_X50_ok']:.1f}%")
            col_b.metric("PPV within limit probability", f"{mc['pct_PPV_ok']:.1f}%")

        with tab5:
            st.subheader("📥 Export Results")
            st.markdown("Download the complete analysis as an **Excel workbook** with 4 sheets.")
            xlsx_bytes = export_excel(p, kuz, ppv_res, spc, cap, mc, violations)
            st.download_button(
                label="⬇️ Download Excel Report (.xlsx)",
                data=xlsx_bytes,
                file_name="SPC_Voladura_Analysis.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

if __name__ == "__main__":
    main()
